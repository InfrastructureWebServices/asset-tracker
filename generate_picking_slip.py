from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment
from os import path
import re
reference_pattern = re.compile(r'([A-Z]+)([0-9]+)')

directory = path.dirname(__file__)

wb = load_workbook(path.join(directory, 'data', 'template.xlsx'))


all_borders = Border(
    left=Side(border_style='thin', color='00000000'),
    right=Side(border_style='thin', color='00000000'),
    bottom=Side(border_style='thin', color='00000000'),
    top=Side(border_style='thin', color='00000000')
    )

center_wrapped = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

import json
with open('picking_slip.json', 'r') as f:
    pick_data = json.loads(f.read())

ws = wb['Form']

defined_names = wb.defined_names

asset_table = ws.tables['asset_table']

insert_row_id = int(defined_names['insert_row_after'].value)

row_count = 0
for i in range(0, len(pick_data)-1):
    ws.insert_rows(insert_row_id, amount=1)
    row_count += 1


def shift_row_ref(ref, shift):
    ref = reference_pattern.findall(ref)
    ref = ref[0][0] + str(int(ref[0][1]) + shift)
    return ref

[from_ref, to_ref] = asset_table.ref.split(':')
asset_table.ref = from_ref + ":" + shift_row_ref(to_ref, row_count)
[from_ref, to_ref] = asset_table.ref.split(':')
data_range = ws[shift_row_ref(from_ref, 1) + ":" + to_ref]

def set_cell(cell, value):
    if value != None:
        cell.value = value
    cell.border = all_borders
    cell.alignment = center_wrapped
    return cell

domain = "http://localhost:5000/"

def set_linked_cell(cell, value):
    cell = set_cell(cell, value)
    cell.hyperlink = "%s/assets/%s" % (domain, value)
    cell.style = 'Hyperlink'
    return cell

def set_pick_item(row, item):
    set_linked_cell(row[0], item.get('id'))
    set_cell(row[1], item.get('description'))
    set_cell(row[2], item.get('manufacturer'))
    set_cell(row[3], item.get('part_number'))
    set_cell(row[4], item.get('serial_number'))
    set_cell(row[5], 1)

row_count = 0
for item in pick_data:
    set_pick_item(data_range[row_count], item)    
    row_count += 1

for row in ws[asset_table.ref]:
    for cell in row:
        cell.border = all_borders

# shift or remove defined name references?


wb.save(path.join(directory, 'output', 'output.xlsx'))

print('ok')
